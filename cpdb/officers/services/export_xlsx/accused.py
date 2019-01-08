from openpyxl import Workbook
from rest_framework.serializers import (
    Serializer, CharField, DateField, BooleanField, IntegerField, FloatField, NullBooleanField, DateTimeField
)

from data.models import OfficerAllegation, Allegation, PoliceWitness, Area, Officer, Victim


class OfficerAllegationExcelSerializer(Serializer):
    # from allegation
    crid = CharField(source='allegation.crid')
    address = CharField(allow_blank=True, source='allegation.address')
    old_complaint_address = CharField(allow_null=True, source='allegation.old_complaint_address')
    incident_date = DateTimeField(format='%Y-%m-%d', allow_null=True, source='allegation.incident_date')
    is_officer_complaint = BooleanField(source='allegation.is_officer_complaint')
    beat = CharField(source='allegation.beat.name', allow_null=True)
    coaccused_count = IntegerField(allow_null=True, source='allegation.coaccused_count')

    category = CharField(allow_null=True)
    subcategory = CharField(allow_null=True)
    start_date = DateField(allow_null=True)
    end_date = DateField(allow_null=True)
    recc_finding = CharField(allow_blank=True, source='recc_finding_display')
    recc_outcome = CharField(allow_blank=True)
    final_finding = CharField(allow_blank=True, source='final_finding_display')
    final_outcome = CharField(allow_blank=True)
    disciplined = NullBooleanField()


class AreaExcelSerializer(Serializer):
    name = CharField()
    area_type = CharField()
    median_income = CharField(allow_null=True)
    commander = CharField(source='commander.full_name', allow_null=True)
    alderman = CharField(allow_null=True)
    police_hq = CharField(source='police_hq.name', allow_null=True)


class PoliceWitnessExcelSerializer(Serializer):
    crid = CharField(source='allegation.crid')
    name = CharField(source='officer.full_name')
    gender = CharField(source='officer.gender_display', allow_blank=True)
    race = CharField(source='officer.race')
    appointed_date = DateField(allow_null=True, source='officer.appointed_date')
    resignation_date = DateField(allow_null=True, source='officer.resignation_date')
    rank = CharField(allow_blank=True, source='officer.rank')
    birth_year = IntegerField(allow_null=True, source='officer.birth_year')
    active = CharField(source='officer.active')

    complaint_percentile = FloatField(allow_null=True, source='officer.complaint_percentile')
    civilian_allegation_percentile = FloatField(allow_null=True, source='officer.civilian_allegation_percentile')
    internal_allegation_percentile = FloatField(allow_null=True, source='officer.internal_allegation_percentile')
    trr_percentile = FloatField(allow_null=True, source='officer.trr_percentile')
    honorable_mention_percentile = FloatField(allow_null=True, source='officer.honorable_mention_percentile')
    allegation_count = IntegerField(allow_null=True, source='officer.allegation_count')
    sustained_count = IntegerField(allow_null=True, source='officer.sustained_count')
    honorable_mention_count = IntegerField(allow_null=True, source='officer.honorable_mention_count')
    unsustained_count = IntegerField(allow_null=True, source='officer.unsustained_count')
    discipline_count = IntegerField(allow_null=True, source='officer.discipline_count')
    civilian_compliment_count = IntegerField(allow_null=True, source='officer.civilian_compliment_count')
    trr_count = IntegerField(allow_null=True, source='officer.trr_count')
    major_award_count = IntegerField(allow_null=True, source='officer.major_award_count')
    current_badge = CharField(allow_null=True, source='officer.current_badge')
    last_unit = CharField(allow_null=True, source='officer.last_unit.unit_name')
    current_salary = IntegerField(allow_null=True, source='officer.current_salary')


class VictimExcelSerializer(Serializer):
    crid = CharField(source='allegation.crid')
    gender = CharField(source='gender_display', allow_blank=True)
    race = CharField()
    birth_year = IntegerField(allow_null=True)


class OfficerExcelWriter(object):
    def __init__(self, officer):
        self.officer = officer
        self.wb = Workbook()

    def write_sheet(self, ws, rows):
        if len(rows) > 0:
            for column_idx, column_name in enumerate(rows[0].keys()):
                ws.cell(row=1, column=column_idx + 1, value=column_name)

            for row_idx, row in enumerate(rows):
                for column_idx, value in enumerate(row.values()):
                    ws.cell(row=row_idx + 2, column=column_idx + 1, value=value)

    @property
    def file_name(self):
        raise NotImplementedError

    def export_xlsx(self):
        raise NotImplementedError

    def save(self):
        self.wb.remove(self.wb['Sheet'])
        self.wb.save(self.file_name)


class AccusedExcelWriter(OfficerExcelWriter):
    @property
    def file_name(self):
        return f'accused_{self.officer.id}.xlsx'

    def write_allegation_sheet(self):
        ws = self.wb.create_sheet('Allegation', 0)
        officer_allegations = OfficerAllegation.objects.filter(
            officer=self.officer
        ).select_related('allegation')
        rows = OfficerAllegationExcelSerializer(officer_allegations, many=True).data
        self.write_sheet(ws, rows)

    def write_police_witnesses_sheet(self):
        ws = self.wb.create_sheet('Police Witness', 1)
        police_witnesses = PoliceWitness.objects.filter(
            allegation__officerallegation__officer=self.officer
        ).distinct().select_related('officer', 'allegation')
        rows = PoliceWitnessExcelSerializer(police_witnesses, many=True).data
        self.write_sheet(ws, rows)

    def write_beat_sheet(self):
        ws = self.wb.create_sheet('Beat', 2)
        beat_ids = Allegation.objects.filter(
            officerallegation__officer=self.officer
        ).distinct().values_list('beat', flat=True).distinct()
        beats = Area.objects.filter(id__in=list(beat_ids)).select_related('police_hq')
        rows = AreaExcelSerializer(beats, many=True).data
        self.write_sheet(ws, rows)

    def write_victim_sheet(self):
        ws = self.wb.create_sheet('Victim', 3)
        victims = Victim.objects.filter(
            allegation__officerallegation__officer=self.officer
        ).select_related('allegation')
        rows = VictimExcelSerializer(victims, many=True).data
        self.write_sheet(ws, rows)

    def export_xlsx(self):
        self.write_allegation_sheet()
        self.write_police_witnesses_sheet()
        self.write_beat_sheet()
        self.write_victim_sheet()
        self.save()
