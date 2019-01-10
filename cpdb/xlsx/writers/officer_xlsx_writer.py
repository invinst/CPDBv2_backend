import os

from openpyxl import Workbook


class OfficerXlsxWriter(object):
    def __init__(self, officer, out_dir):
        self.officer = officer
        self.out_dir = out_dir
        self.wb = Workbook()

    def write_sheet(self, ws, queryset, serializer_klass):
        rows = serializer_klass(queryset, many=True).data
        fields = list(serializer_klass().fields)

        for column_idx, column_name in enumerate(fields):
            ws.cell(row=1, column=column_idx + 1, value=column_name)

        for row_idx, row in enumerate(rows):
            for column_idx, value in enumerate(row.values()):
                ws.cell(row=row_idx + 2, column=column_idx + 1, value=value)

    @property
    def file_name(self):
        raise NotImplementedError

    def export_xlsx(self):
        raise NotImplementedError

    def save(self):
        if not os.path.exists(self.out_dir):
            os.mkdir(self.out_dir)

        self.wb.remove(self.wb['Sheet'])
        self.wb.save(f'{self.out_dir}/{self.file_name}')
