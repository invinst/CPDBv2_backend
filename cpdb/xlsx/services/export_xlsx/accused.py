from openpyxl import Workbook
from rest_framework.serializers import (
    Serializer, CharField, DateField, IntegerField, FloatField
)

from data.models import OfficerAllegation, Allegation, PoliceWitness, Area, Victim
from xlsx.serializers.accused import OfficerAllegationExcelSerializer


class OfficerExcelWriter(object):
    def __init__(self, officer):
        self.officer = officer
        self.wb = Workbook()

    def write_sheet(self, ws, rows):
        if len(rows) > 0:
            for column_idx, column_name in enumerate(rows[0].keys()):
                ws.cell(row=1, column=column_idx + 1, value=column_name)

            for row_idx, row in enumerate(rows):
                for column_idx, value in enumerate(row.values()):
                    ws.cell(row=row_idx + 2, column=column_idx + 1, value=value)

    @property
    def file_name(self):
        raise NotImplementedError

    def export_xlsx(self):
        raise NotImplementedError

    def save(self):
        self.wb.remove(self.wb['Sheet'])
        self.wb.save(self.file_name)


class AccusedExcelWriter(OfficerExcelWriter):
    @property
    def file_name(self):
        return f'accused_{self.officer.id}.xlsx'

    def write_allegation_sheet(self):
        ws = self.wb.create_sheet('Allegation', 0)
        officer_allegations = OfficerAllegation.objects.filter(
            officer=self.officer
        ).select_related('allegation')
        rows = OfficerAllegationExcelSerializer(officer_allegations, many=True).data
        self.write_sheet(ws, rows)

    def write_police_witnesses_sheet(self):
        ws = self.wb.create_sheet('Police Witness', 1)
        police_witnesses = PoliceWitness.objects.filter(
            allegation__officerallegation__officer=self.officer
        ).distinct().select_related('officer', 'allegation')
        rows = PoliceWitnessExcelSerializer(police_witnesses, many=True).data
        self.write_sheet(ws, rows)

    def write_beat_sheet(self):
        ws = self.wb.create_sheet('Beat', 2)
        beat_ids = Allegation.objects.filter(
            officerallegation__officer=self.officer
        ).distinct().values_list('beat', flat=True).distinct()
        beats = Area.objects.filter(id__in=list(beat_ids)).select_related('police_hq')
        rows = AreaExcelSerializer(beats, many=True).data
        self.write_sheet(ws, rows)

    def write_victim_sheet(self):
        ws = self.wb.create_sheet('Victim', 3)
        victims = Victim.objects.filter(
            allegation__officerallegation__officer=self.officer
        ).select_related('allegation')
        rows = VictimExcelSerializer(victims, many=True).data
        self.write_sheet(ws, rows)

    def export_xlsx(self):
        self.write_allegation_sheet()
        self.write_police_witnesses_sheet()
        self.write_beat_sheet()
        self.write_victim_sheet()
        self.save()
